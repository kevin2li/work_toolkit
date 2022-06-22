import glob
import os
import os.path as osp
import traceback
from pathlib import Path

import typer
from loguru import logger
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

app = typer.Typer()

@app.command(name="merge", help="merge multiple xlsx files in the specified directory")
def merge(
    dir: str = typer.Argument(...),
    output_path: str = typer.Option(None, "--output-path", "-o", help="output path"),
):
    try:
        result_wb = Workbook()
        result_ws = result_wb.active
        addHeaderDone = False

        path_list = glob.glob(dir, "*.xlsx")
        for path in path_list:
            wb = load_workbook(path)
            ws = wb.active
            for i, row in enumerate(ws.rows):
                values = [c.value for c in row]
                if i == 0:
                    if not addHeaderDone:
                        result_ws.append(values)
                        addHeaderDone = True
                else:
                    result_ws.append(values)
        if output_path is None:
            output_path = osp.join(os.getcwd(), f"merged.xlsx")

        result_wb.save(output_path)
    except:
        logger.error("合并时发生错误！")
        traceback.print_exc()

if __name__ == "__main__":
    app()
